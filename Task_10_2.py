import openpyxl
wb = openpyxl.load_workbook("Task_10_2.xlsx")
ws = wb.active

ws = wb.copy_worksheet(ws)
wb.active = wb['Sheet Copy']
ws = wb.active

su = 0
for i in range(1, 4):
    print((ws.cell(i, 1).value, ws.cell(i, 2).value)[::-1])
    su += ws.cell(i, 2).value
ws.cell(1, 3).value = su

wb.save("Task_10_2.xlsx")


